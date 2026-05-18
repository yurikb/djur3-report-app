import io
import json
import os
from collections import Counter
from datetime import datetime

import openpyxl


def limpar(texto):
    if texto is None:
        return ""
    return str(texto).strip()


def extrair(caminho_ou_bytes):
    if isinstance(caminho_ou_bytes, (bytes, io.BytesIO)):
        wb = openpyxl.load_workbook(caminho_ou_bytes, data_only=True)
    else:
        wb = openpyxl.load_workbook(caminho_ou_bytes, data_only=True)

    ws = wb["CPJ3C"]
    headers = [c.value for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    dados = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: row[i] for i in range(len(headers))}
        dados.append(d)

    return dados, headers


def resumir_dados(dados):
    tipos = Counter()
    fases = Counter()
    ufs = Counter()
    relevancias = Counter()
    grupos = Counter()
    valores = []
    obs = []

    for d in dados:
        ta = limpar(d.get("Tipo de ação", ""))
        if ta: tipos[ta] += 1
        fa = limpar(d.get("PRO.FAS.Descrição", ""))
        if fa: fases[fa] += 1
        uf = limpar(d.get("UF", ""))
        if uf: ufs[uf] += 1
        rl = limpar(d.get("Relevancia", ""))
        if rl: relevancias[rl] += 1
        gr = limpar(d.get("PRO.GRU.Descrição", ""))
        if gr: grupos[gr] += 1
        val = d.get("PRO.Valor da causa")
        if val: valores.append(float(val))
        txt = limpar(d.get("ATP.Texto", ""))
        if txt and txt not in ("0", "nan", "None", ""):
            obs.append({"PJ": d.get("PJ"), "texto": txt, "tipo": ta})

    return {
        "total": len(dados),
        "valor_total": sum(valores),
        "valor_medio": sum(valores) / len(valores) if valores else 0,
        "tipos": dict(tipos.most_common()),
        "fases": dict(fases.most_common()),
        "ufs": dict(ufs.most_common()),
        "relevancias": dict(relevancias.most_common()),
        "grupos": dict(grupos.most_common()),
        "observacoes": obs,
        "prazos": sorted(
            [
                {
                    "PJ": d.get("PJ"),
                    "processo": limpar(d.get("PRO.Número do processo", "")),
                    "agenda": str(d.get("Agenda", "") or "")[:10],
                    "fatal": str(d.get("Fatal", "") or "")[:10],
                    "fase": limpar(d.get("PRO.FAS.Descrição", "")),
                    "tipo": limpar(d.get("Tipo de ação", "")),
                    "relevancia": limpar(d.get("Relevancia", "")),
                }
                for d in dados if d.get("Agenda")
            ],
            key=lambda x: x["agenda"],
        )[:10],
    }


def gerar_relatorio_template(dados, headers, prompt_extra=""):
    r = resumir_dados(dados)
    linhas = []

    def P(t="", end="\n"):
        linhas.append(str(t) + end)

    P("# Relatório de Processos — DJUR3")
    P()
    P(f"**Total de processos:** {r['total']}")
    P(f"**Valor total:** R$ {r['valor_total']:,.2f}")
    P(f"**Valor médio:** R$ {r['valor_medio']:,.2f}")
    P()

    if prompt_extra.strip():
        P(f"**Objetivo do relatório:** {prompt_extra}")
        P()

    P("## 1. Dashboard Geral")
    P()
    P("| Indicador | Valor |")
    P("|---|---|")
    P(f"| Total de processos | {r['total']} |")
    P(f"| Valor total da causa | R$ {r['valor_total']:,.2f} |")
    P(f"| Média por processo | R$ {r['valor_medio']:,.2f} |")
    P()

    P("### Distribuição por UF")
    P("| UF | Qtd |")
    P("|---|---|")
    for k, v in r["ufs"].items():
        P(f"| {k} | {v} |")
    P()

    P("### Cooperativas")
    P("| Cooperativa | Qtd |")
    P("|---|---|")
    for k, v in r["grupos"].items():
        P(f"| {k} | {v} |")
    P()

    P("## 2. Por Tipo de Ação")
    P()
    P("| # | Tipo | Qtd | % |")
    P("|---|---|---|---|")
    for i, (k, v) in enumerate(r["tipos"].items(), 1):
        P(f"| {i} | {k} | {v} | {v/r['total']*100:.1f}% |")
    P()

    P("## 3. Por Relevância")
    P()
    for k, v in r["relevancias"].items():
        P(f"- **{k}:** {v}")
    P()

    P("## 4. Por Fase Processual")
    P()
    for k, v in r["fases"].items():
        P(f"- **{k}:** {v}")
    P()

    P("## 5. Casos com Observações")
    P()
    if r["observacoes"]:
        for o in r["observacoes"]:
            P(f"### PJ {o['PJ']} — {o['tipo']}")
            P(f"> {o['texto']}")
            P()
    else:
        P("Nenhuma observação registrada.")
        P()

    P("## 6. Próximos Prazos (Top 10)")
    P()
    P("| PJ | Processo | Agenda | Fatal | Fase | Tipo | Relevância |")
    P("|---|---|---|---|---|---|---|")
    for p in r["prazos"]:
        P(f"| {p['PJ']} | {p['processo']} | {p['agenda']} | {p['fatal']} | {p['fase']} | {p['tipo']} | {p['relevancia']} |")
    P()

    return "".join(linhas), r


AGENT_INSTRUCOES = """# Juridic Report Extractor — XLSX

Extrai dados de planilhas do sistema DJUR3 e gera relatórios analíticos em markdown.

## Output esperado
1. Dashboard Geral (total, valores, UFs, cooperativas)
2. Por Tipo de Ação (qtd, valores, exemplos)
3. Por Relevância / Prioridade
4. Por Fase Processual
5. Casos com Observações (ATP.Texto)
6. Prazos (Agenda/Fatal) mais próximos
7. Summary (se solicitado tom específico)

## Regras OBRIGATÓRIAS
1. NUNCA inventar dados, jurisprudências, prazos ou informações que não estão no arquivo.
2. SEMPRE mencionar somente o que está nos dados fornecidos.
3. Corroborar com valores brutos da planilha sempre que possível.
4. Preservar números de processo exatos.
5. Valores da causa em R$.
6. Se o prompt pedir tom de "minimizar erro" ou "destacar condução correta", ajustar a narrativa."""


def gerar_relatorio_ia(dados, headers, prompt_extra=""):
    """Gera relatório via API compatível OpenAI (OpenRouter / DeepSeek)."""
    r = resumir_dados(dados)

    system = AGENT_INSTRUCOES
    if prompt_extra.strip():
        system += f"\n\n## Configuração adicional do usuário\n{prompt_extra}"

    user_msg = f"""Analise estes {r['total']} processos do DJUR3 e gere um relatório em markdown seguindo as instruções do sistema.

## Dados Consolidados
- Total: {r['total']} processos
- Valor total: R$ {r['valor_total']:,.2f}
- Valor médio: R$ {r['valor_medio']:,.2f}

### Distribuição por UF
{json.dumps(r['ufs'], ensure_ascii=False, indent=2)}

### Por Cooperativa
{json.dumps(r['grupos'], ensure_ascii=False, indent=2)}

### Tipos de Ação
{json.dumps(r['tipos'], ensure_ascii=False, indent=2)}

### Fases Processuais
{json.dumps(r['fases'], ensure_ascii=False, indent=2)}

### Relevância
{json.dumps(r['relevancias'], ensure_ascii=False, indent=2)}

### Observações Relevantes
{json.dumps(r['observacoes'], ensure_ascii=False, indent=2)}

### Próximos Prazos
{json.dumps(r['prazos'], ensure_ascii=False, indent=2)}
"""

    api_url = os.environ.get("AI_API_URL", "https://openrouter.ai/api/v1/chat/completions")
    api_key = os.environ.get("AI_API_KEY", "") or os.environ.get("OPENROUTER_API_KEY", "") or os.environ.get("OPENAI_API_KEY", "")
    model = os.environ.get("AI_MODEL", "deepseek/deepseek-v4-flash")

    if not api_key:
        return None, "API key não configurada. Configure OPENROUTER_API_KEY, AI_API_KEY ou OPENAI_API_KEY."

    headers_api = {"Content-Type": "application/json"}
    if "openrouter" in api_url:
        headers_api["Authorization"] = f"Bearer {api_key}"
        headers_api["HTTP-Referer"] = "https://djur3-report-app.streamlit.app"
    else:
        headers_api["Authorization"] = f"Bearer {api_key}"

    import httpx

    try:
        resp = httpx.post(
            api_url,
            headers=headers_api,
            json={
                "model": model,
                "messages": [
                    {"role": "system", "content": system},
                    {"role": "user", "content": user_msg},
                ],
                "temperature": 0.3,
                "max_tokens": 4096,
            },
            timeout=120,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"]
        return content, None
    except Exception as e:
        return None, str(e)
